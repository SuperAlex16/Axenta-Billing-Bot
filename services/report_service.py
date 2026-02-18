"""Сервис генерации отчётов (Excel и PDF) для статистики расчётов"""
import calendar
from collections import defaultdict
from io import BytesIO
from datetime import date, datetime
from typing import List, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.fonts import addMapping

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
from utils.constants import MONTHS_RU
from utils.logger import setup_logger

logger = setup_logger(__name__)

# Регистрация кириллического шрифта для PDF
_font_registered = False


def _register_cyrillic_font():
    """Регистрация шрифта с поддержкой кириллицы для PDF (regular + bold)"""
    global _font_registered
    if _font_registered:
        return

    # Пары (regular, bold)
    font_pairs = [
        ('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
         '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'),
        ('/usr/share/fonts/TTF/DejaVuSans.ttf',
         '/usr/share/fonts/TTF/DejaVuSans-Bold.ttf'),
        ('C:/Windows/Fonts/arial.ttf',
         'C:/Windows/Fonts/arialbd.ttf'),
        ('C:/Windows/Fonts/calibri.ttf',
         'C:/Windows/Fonts/calibrib.ttf'),
    ]

    for regular_path, bold_path in font_pairs:
        try:
            if Path(regular_path).exists():
                pdfmetrics.registerFont(TTFont('CyrillicFont', regular_path))
                if Path(bold_path).exists():
                    pdfmetrics.registerFont(TTFont('CyrillicFont-Bold', bold_path))
                else:
                    pdfmetrics.registerFont(TTFont('CyrillicFont-Bold', regular_path))
                addMapping('CyrillicFont', 0, 0, 'CyrillicFont')       # normal
                addMapping('CyrillicFont', 1, 0, 'CyrillicFont-Bold')  # bold
                _font_registered = True
                logger.info(f"PDF шрифт зарегистрирован: {regular_path}")
                return
        except Exception:
            continue

    logger.warning("Кириллический шрифт для PDF не найден, используется Helvetica")


def _get_period_dates(year: int, month: int) -> List[date]:
    """Возвращает список дат для месячного отчёта."""
    days_in_month = calendar.monthrange(year, month)[1]
    return [date(year, month, d) for d in range(1, days_in_month + 1)]


def _get_period_label(year: int, month: Optional[int]) -> str:
    """Возвращает текстовое описание периода"""
    if month is not None:
        return f"{MONTHS_RU.get(month, str(month))} {year}"
    return str(year)


def _aggregate_monthly_data(year: int, charges_data: List[Dict],
                            payments_data: Dict[date, float]) -> List[Dict]:
    """
    Агрегация данных по месяцам для годового отчёта.

    Returns:
        Список dict: month, month_name, objects, tariff, charge, payment
    """
    today = date.today()

    # Группируем списания по месяцам
    charges_by_month = defaultdict(list)
    for item in charges_data:
        charges_by_month[item['date'].month].append(item)

    # Группируем поступления по месяцам
    payments_by_month = defaultdict(float)
    for dt, amount in payments_data.items():
        payments_by_month[dt.month] += amount

    max_month = 12 if year < today.year else today.month

    monthly = []
    for m in range(1, max_month + 1):
        month_charges = charges_by_month.get(m, [])

        if month_charges:
            avg_objects = round(sum(c['objects'] for c in month_charges) / len(month_charges))
            avg_tariff = sum(c['tariff'] for c in month_charges) / len(month_charges)
            total_charge = sum(c['charge'] for c in month_charges)
        else:
            avg_objects = 0
            avg_tariff = 0.0
            total_charge = 0.0

        total_payment = payments_by_month.get(m, 0.0)

        monthly.append({
            'month': m,
            'month_name': MONTHS_RU[m],
            'objects': avg_objects,
            'tariff': round(avg_tariff, 2),
            'charge': total_charge,
            'payment': total_payment,
        })

    return monthly


def generate_excel(account_login: str, year: int, month: Optional[int],
                   charges_data: List[Dict], payments_data: Dict[date, float],
                   initial_balance: float, organization: str = '') -> BytesIO:
    """
    Генерация Excel-отчёта «Акт сверки расчётов за период».

    Args:
        account_login: логин аккаунта
        year: год
        month: месяц (1-12) или None для годового отчёта
        charges_data: список dict {'date': date, 'objects': int, 'tariff': float, 'charge': float}
        payments_data: dict {date: float} — суммарные поступления по датам
        initial_balance: начальный баланс на начало периода
        organization: название организации из листа «Данные»

    Returns:
        BytesIO объект с xlsx файлом
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Акт сверки"

    # Стили
    header_font = Font(name='Calibri', size=14, bold=True)
    sub_font = Font(name='Calibri', size=11)
    label_font = Font(name='Calibri', size=11, bold=True)
    table_header_font = Font(name='Calibri', size=11, bold=True)
    data_font = Font(name='Calibri', size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовочная часть
    ws.merge_cells('A1:F1')
    ws['A1'] = 'Акт сверки расчётов за период'
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A3'] = 'Аккаунт:'
    ws['A3'].font = label_font
    ws['B3'] = account_login
    ws['B3'].font = sub_font

    ws['A4'] = 'Организация:'
    ws['A4'].font = label_font
    ws['B4'] = organization
    ws['B4'].font = sub_font

    period_label = _get_period_label(year, month)
    period_field = 'Год:' if month is None else 'Период:'
    ws['A6'] = period_field
    ws['A6'].font = label_font
    ws['B6'] = period_label
    ws['B6'].font = sub_font

    # Заголовки таблицы (строка 8)
    table_headers = [
        'Дата',
        'Кол-во Активных\nобъектов, шт',
        'Тариф за 1 объект\nв сутки, руб',
        'Сумма Списания,\nруб',
        'Сумма Поступлений,\nруб',
        'Остаток баланса\nна конец дня, руб'
    ]

    for col, header_text in enumerate(table_headers, 1):
        cell = ws.cell(row=8, column=col, value=header_text)
        cell.font = table_header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Подготовка общих данных
    running_balance = initial_balance
    total_charge_sum = 0.0
    total_payment_sum = 0.0
    row_num = 9

    if month is None:
        # === Годовой отчёт: консолидированные данные по месяцам ===
        monthly_data = _aggregate_monthly_data(year, charges_data, payments_data)

        for item in monthly_data:
            m_charge = item['charge']
            m_payment = item['payment']
            m_objects = item['objects']
            m_tariff = item['tariff']

            running_balance += m_payment + m_charge
            total_charge_sum += m_charge
            total_payment_sum += m_payment

            # Месяц
            cell = ws.cell(row=row_num, column=1, value=item['month_name'])
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

            # Активные объекты
            cell = ws.cell(row=row_num, column=2, value=m_objects if m_objects else '')
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if m_objects:
                cell.number_format = '0'

            # Тариф
            cell = ws.cell(row=row_num, column=3, value=m_tariff if m_tariff else '')
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right')
            if m_tariff:
                cell.number_format = '0.00'

            # Сумма списания
            cell = ws.cell(row=row_num, column=4, value=round(m_charge, 2) if m_charge != 0.0 else '')
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right')
            if m_charge != 0.0:
                cell.number_format = '0.00'

            # Сумма поступлений
            cell = ws.cell(row=row_num, column=5, value=round(m_payment, 2) if m_payment else '')
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right')
            if m_payment:
                cell.number_format = '0.00'

            # Остаток баланса
            cell = ws.cell(row=row_num, column=6, value=round(running_balance, 2))
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right')
            cell.number_format = '0.00'

            row_num += 1
    else:
        # === Месячный отчёт: только дни с движениями ===
        charges_by_date = {item['date']: item for item in charges_data}
        period_dates = _get_period_dates(year, month)

        for current_date in period_dates:
            charge_item = charges_by_date.get(current_date)
            payment_amount = payments_data.get(current_date)

            # Пропускаем дни без движений
            if not charge_item and not payment_amount:
                continue

            day_charge = charge_item['charge'] if charge_item else 0.0
            day_objects = charge_item['objects'] if charge_item else ''
            day_tariff = charge_item['tariff'] if charge_item else ''
            day_payment = payment_amount if payment_amount else None

            running_balance = running_balance + (day_payment or 0.0) + day_charge
            total_charge_sum += day_charge
            total_payment_sum += (day_payment or 0.0)

            # Дата
            cell = ws.cell(row=row_num, column=1, value=current_date.strftime('%d.%m.%Y'))
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

            # Активные объекты
            cell = ws.cell(row=row_num, column=2, value=day_objects if day_objects != '' else '')
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if isinstance(day_objects, int):
                cell.number_format = '0'

            # Тариф
            cell = ws.cell(row=row_num, column=3, value=day_tariff if day_tariff != '' else '')
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right')
            if isinstance(day_tariff, float):
                cell.number_format = '0.00'

            # Сумма списания
            cell = ws.cell(row=row_num, column=4, value=day_charge if day_charge != 0.0 else '')
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right')
            if day_charge != 0.0:
                cell.number_format = '0.00'

            # Сумма поступлений
            cell = ws.cell(row=row_num, column=5, value=day_payment if day_payment else '')
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right')
            if day_payment:
                cell.number_format = '0.00'

            # Остаток баланса
            cell = ws.cell(row=row_num, column=6, value=round(running_balance, 2))
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right')
            cell.number_format = '0.00'

            row_num += 1

    # Итоговая строка
    row_num += 1
    ws.cell(row=row_num, column=1, value='Итого за период:').font = Font(name='Calibri', size=11, bold=True)
    row_num += 1
    ws.cell(row=row_num, column=1, value='Сумма списаний:').font = sub_font
    cell = ws.cell(row=row_num, column=2, value=round(total_charge_sum, 2))
    cell.font = sub_font
    cell.number_format = '0.00'
    ws.cell(row=row_num, column=3, value='руб').font = sub_font

    row_num += 1
    ws.cell(row=row_num, column=1, value='Сумма поступлений:').font = sub_font
    cell = ws.cell(row=row_num, column=2, value=round(total_payment_sum, 2))
    cell.font = sub_font
    cell.number_format = '0.00'
    ws.cell(row=row_num, column=3, value='руб').font = sub_font

    # Автоширина столбцов
    column_widths = [21, 18, 18, 17, 18, 22]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Настройка печати: вертикальная А4, всё на одной странице
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.4, right=0.4, top=0.5, bottom=0.5,
        header=0.3, footer=0.3
    )

    # Сохранение в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info(f"Excel-отчёт сформирован для {account_login} за {period_label}")
    return output


def generate_pdf(account_login: str, year: int, month: Optional[int],
                 charges_data: List[Dict], payments_data: Dict[date, float],
                 initial_balance: float, organization: str = '') -> BytesIO:
    """
    Генерация PDF-отчёта «Акт сверки расчётов за период».

    Args:
        account_login: логин аккаунта
        year: год
        month: месяц (1-12) или None для годового отчёта
        charges_data: список dict {'date': date, 'objects': int, 'tariff': float, 'charge': float}
        payments_data: dict {date: float} — суммарные поступления по датам
        initial_balance: начальный баланс на начало периода
        organization: название организации из листа «Данные»

    Returns:
        BytesIO объект с pdf файлом
    """
    _register_cyrillic_font()

    font_name = 'CyrillicFont' if _font_registered else 'Helvetica'

    output = BytesIO()

    page_size = A4

    doc = SimpleDocTemplate(
        output,
        pagesize=page_size,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm
    )

    elements = []

    # Стили
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleRu', parent=styles['Title'],
        fontName=font_name, fontSize=14, leading=18
    )
    normal_style = ParagraphStyle(
        'NormalRu', parent=styles['Normal'],
        fontName=font_name, fontSize=10, leading=14
    )
    bold_style = ParagraphStyle(
        'BoldNormalRu', parent=normal_style,
        fontName=font_name, fontSize=10, leading=14
    )

    period_label = _get_period_label(year, month)

    # Заголовок
    elements.append(Paragraph('<b>Акт сверки расчётов за период</b>', title_style))
    elements.append(Spacer(1, 6 * mm))

    elements.append(Paragraph(f'<b>Аккаунт:</b> {account_login}', bold_style))
    elements.append(Spacer(1, 2 * mm))
    elements.append(Paragraph(f'<b>Организация:</b> {organization}', bold_style))
    elements.append(Spacer(1, 5 * mm))
    period_field = 'Год' if month is None else 'Период'
    elements.append(Paragraph(f'<b>{period_field}:</b> {period_label}', bold_style))
    elements.append(Spacer(1, 6 * mm))

    # Заголовки таблицы
    table_data = [[
        'Дата',
        'Кол-во Актив.\nобъектов, шт',
        'Тариф за 1\nобъект, руб',
        'Сумма\nСписания, руб',
        'Сумма\nПоступл., руб',
        'Остаток\nбаланса, руб'
    ]]

    running_balance = initial_balance
    total_charge_sum = 0.0
    total_payment_sum = 0.0

    if month is None:
        # === Годовой отчёт: консолидированные данные по месяцам ===
        monthly_data = _aggregate_monthly_data(year, charges_data, payments_data)

        for item in monthly_data:
            m_charge = item['charge']
            m_payment = item['payment']

            running_balance += m_payment + m_charge
            total_charge_sum += m_charge
            total_payment_sum += m_payment

            table_data.append([
                item['month_name'],
                str(item['objects']) if item['objects'] else '',
                f"{item['tariff']:.2f}" if item['tariff'] else '',
                f"{m_charge:.2f}" if m_charge != 0.0 else '',
                f"{m_payment:.2f}" if m_payment else '',
                f"{running_balance:.2f}"
            ])
    else:
        # === Месячный отчёт: только дни с движениями ===
        charges_by_date = {item['date']: item for item in charges_data}
        period_dates = _get_period_dates(year, month)

        for current_date in period_dates:
            charge_item = charges_by_date.get(current_date)
            payment_amount = payments_data.get(current_date)

            # Пропускаем дни без движений
            if not charge_item and not payment_amount:
                continue

            day_charge = charge_item['charge'] if charge_item else 0.0
            day_objects = str(charge_item['objects']) if charge_item else ''
            day_tariff = f"{charge_item['tariff']:.2f}" if charge_item else ''
            day_payment = f"{payment_amount:.2f}" if payment_amount else ''
            day_charge_str = f"{day_charge:.2f}" if day_charge != 0.0 else ''

            running_balance = running_balance + (payment_amount or 0.0) + day_charge
            total_charge_sum += day_charge
            total_payment_sum += (payment_amount or 0.0)

            table_data.append([
                current_date.strftime('%d.%m.%Y'),
                day_objects,
                day_tariff,
                day_charge_str,
                day_payment,
                f"{running_balance:.2f}"
            ])

    # Создание таблицы — растягиваем на всю ширину страницы
    available_width = A4[0] - 15 * mm - 15 * mm  # ширина страницы минус поля
    col_widths = [
        available_width * 0.13,  # Дата
        available_width * 0.17,  # Кол-во объектов
        available_width * 0.17,  # Тариф
        available_width * 0.17,  # Списания
        available_width * 0.17,  # Поступления
        available_width * 0.19,  # Остаток баланса
    ]

    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('LEADING', (0, 0), (-1, -1), 10),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 6 * mm))

    # Итоги
    elements.append(Paragraph('<b>Итого за период:</b>', bold_style))
    elements.append(Paragraph(
        f'  Сумма списаний: {total_charge_sum:.2f} руб', normal_style
    ))
    elements.append(Paragraph(
        f'  Сумма поступлений: {total_payment_sum:.2f} руб', normal_style
    ))
    doc.build(elements)
    output.seek(0)

    logger.info(f"PDF-отчёт сформирован для {account_login} за {period_label}")
    return output
